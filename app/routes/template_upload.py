"""
Template Upload Route
======================
POST endpoint that parses a populated Standard Data Template (Excel)
and auto-fills KPI scores on Screen 3F. The consultant uploads the
file returned by the client's data analyst.

Reads the "Data Template" sheet, matches KPI names to DB records,
and creates/updates KPIScore + KPIAnchor records.
"""

from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import JSONResponse, RedirectResponse
from app.database import SessionLocal
from app.models import Discovery, Process, Parameter, KPI, KPIScore, KPIAnchor
from app.auth import decode_access_token
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

router = APIRouter()


def require_auth(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        return None
    payload = decode_access_token(token)
    return payload


@router.post("/discovery/{discovery_id}/upload-template")
async def upload_template(request: Request, discovery_id: str, file: UploadFile = File(...)):
    user = require_auth(request)
    if not user:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    db = SessionLocal()
    try:
        discovery = db.query(Discovery).filter(Discovery.id == discovery_id).first()
        if not discovery:
            return JSONResponse({"error": "Discovery not found"}, status_code=404)

        process = db.query(Process).filter(Process.discovery_id == discovery.id).first()
        if not process:
            return JSONResponse({"error": "Process not found"}, status_code=404)

        # Read uploaded Excel
        contents = await file.read()
        wb = load_workbook(BytesIO(contents), read_only=False, data_only=True)

        # Find the Data Template sheet
        if "Data Template" not in wb.sheetnames:
            return JSONResponse({"error": "Sheet 'Data Template' not found in uploaded file"}, status_code=400)

        ws = wb["Data Template"]

        # Build KPI lookup by name (case-insensitive)
        parameters = db.query(Parameter).filter(Parameter.process_id == process.id).all()
        kpi_lookup = {}
        for param in parameters:
            kpis = db.query(KPI).filter(KPI.parameter_id == param.id).all()
            for kpi in kpis:
                kpi_lookup[kpi.name.strip().lower()] = kpi

        # Parse rows (skip header row 1)
        updated = 0
        skipped = 0
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            # Columns: Parameter, KPI Name, Formula, Unit, Measurement Period, Raw Value, Evidence Source, Notes
            kpi_name_cell = row[1].value if len(row) > 1 else None
            raw_value_cell = row[5].value if len(row) > 5 else None
            evidence_cell = row[6].value if len(row) > 6 else None
            notes_cell = row[7].value if len(row) > 7 else None

            if not kpi_name_cell:
                continue

            kpi_name_clean = str(kpi_name_cell).strip().lower()
            kpi = kpi_lookup.get(kpi_name_clean)

            if not kpi:
                errors.append("KPI not found: " + str(kpi_name_cell))
                skipped += 1
                continue

            if raw_value_cell is None or str(raw_value_cell).strip() == "" or str(raw_value_cell).strip().upper() == "N/A":
                skipped += 1
                continue

            try:
                raw_value = float(raw_value_cell)
            except (ValueError, TypeError):
                errors.append("Invalid value for " + str(kpi_name_cell) + ": " + str(raw_value_cell))
                skipped += 1
                continue

            # Get anchors (best/worst) from existing KPIAnchor
            anchor = db.query(KPIAnchor).filter(KPIAnchor.kpi_id == kpi.id).first()

            if anchor:
                # Compute normalized score
                best = anchor.best_value
                worst = anchor.worst_value
                if best == worst:
                    normalized = 50.0
                else:
                    normalized = (raw_value - worst) / (best - worst) * 100.0
                    normalized = max(0.0, min(100.0, round(normalized, 1)))
            else:
                # No anchors yet: store raw value as score (0-100 assumed)
                # Consultant will need to set anchors on Screen 3F
                normalized = max(0.0, min(100.0, round(raw_value, 1)))

            # Build evidence text
            evidence_parts = [str(raw_value)]
            if evidence_cell:
                evidence_parts.append(str(evidence_cell))
            if notes_cell:
                evidence_parts.append(str(notes_cell))
            evidence_text = " | ".join(evidence_parts)

            # Save/update score
            existing_score = db.query(KPIScore).filter(
                KPIScore.kpi_id == kpi.id,
                KPIScore.measurement_label == "discovery_baseline"
            ).first()

            if existing_score:
                existing_score.score = normalized
                existing_score.evidence_note = evidence_text
                existing_score.scored_at = datetime.utcnow()
            else:
                new_score = KPIScore(
                    kpi_id=kpi.id,
                    score=normalized,
                    measurement_label="discovery_baseline",
                    evidence_note=evidence_text,
                )
                db.add(new_score)

            updated += 1

        discovery.status = "scored"
        db.commit()

        return JSONResponse({
            "success": True,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:5],
        })

    except Exception as e:
        db.rollback()
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        db.close()
