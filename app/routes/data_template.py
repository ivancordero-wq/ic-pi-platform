"""
Standard Data Template Generator
=================================
Generates the Excel template that the client's data analyst fills
with raw KPI values. Pre-populated with KPI names, formulas, units.
Includes an Instructions sheet explaining what to fill.

Download route: GET /discovery/{id}/data-template
"""

from fastapi import APIRouter, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from app.database import SessionLocal
from app.models import Discovery, Process, Parameter, KPI, Client
from app.auth import decode_access_token
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

router = APIRouter()


def require_auth(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        return None
    payload = decode_access_token(token)
    return payload


def create_instructions_sheet(wb, client_name, process_name):
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = "4472C4"

    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    body_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)

    ws.column_dimensions["A"].width = 80

    rows = [
        ("IC-Pi Standard Data Template", title_font),
        ("", body_font),
        ("Client: " + client_name, bold_font),
        ("Process: " + process_name, bold_font),
        ("", body_font),
        ("PURPOSE", title_font),
        ("This template contains the KPIs identified during the IC-Pi Discovery Phase 1.", body_font),
        ("Your task is to provide the RAW VALUE for each KPI using the formula shown.", body_font),
        ("", body_font),
        ("INSTRUCTIONS", title_font),
        ("1. Go to the 'Data Template' sheet (next tab).", body_font),
        ("2. For each KPI row, read the Formula column to understand what to compute.", body_font),
        ("3. Enter the raw computed value in the 'Raw Value' column.", body_font),
        ("4. Use the UNIT shown (days, %, $, ratio, etc.). Do NOT normalize or scale.", body_font),
        ("5. Fill in the Measurement Period (e.g., 'Q3 2026', 'Aug 2026').", body_font),
        ("6. Fill in Evidence Source (which system or report the data came from).", body_font),
        ("7. Use the Notes column for any caveats, approximations, or data quality flags.", body_font),
        ("8. Return the completed file to your IC-Pi consultant.", body_font),
        ("", body_font),
        ("IMPORTANT NOTES", title_font),
        ("- Do NOT modify the Parameter, KPI Name, Formula, or Unit columns.", body_font),
        ("- Do NOT add or remove rows.", body_font),
        ("- If a KPI cannot be measured, enter 'N/A' in Raw Value and explain in Notes.", body_font),
        ("- IC-Pi will handle all normalization and scoring internally.", body_font),
        ("- This same template will be re-used for each measurement cycle (Stewardship).", body_font),
    ]

    for i, (text, font) in enumerate(rows, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font


def create_data_sheet(wb, kpi_rows):
    ws = wb.create_sheet("Data Template")
    ws.sheet_properties.tabColor = "00B0F0"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    locked_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    locked_font = Font(name="Calibri", size=10, color="666666")
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    body_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Parameter",
        "KPI Name",
        "Formula",
        "Unit",
        "Measurement Period",
        "Raw Value",
        "Evidence Source",
        "Notes",
    ]

    col_widths = [25, 30, 50, 10, 20, 15, 30, 30]

    # Write headers
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write KPI rows
    for row_idx, kpi in enumerate(kpi_rows, 2):
        # Pre-filled columns (locked, grey background)
        for col, key in enumerate(["parameter_name", "kpi_name", "formula", "unit"], 1):
            cell = ws.cell(row=row_idx, column=col, value=kpi.get(key, ""))
            cell.font = locked_font
            cell.fill = locked_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

        # Input columns (yellow background, editable)
        for col in range(5, 9):
            cell = ws.cell(row=row_idx, column=col, value="")
            cell.font = body_font
            cell.fill = input_fill
            cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"


@router.get("/discovery/{discovery_id}/data-template")
async def download_data_template(request: Request, discovery_id: str):
    user = require_auth(request)
    if not user:
        return RedirectResponse(url="/", status_code=302)

    db = SessionLocal()
    try:
        discovery = db.query(Discovery).filter(Discovery.id == discovery_id).first()
        if not discovery:
            return RedirectResponse(url="/dashboard", status_code=302)

        process = db.query(Process).filter(Process.discovery_id == discovery.id).first()
        if not process:
            return RedirectResponse(url="/dashboard", status_code=302)

        client = db.query(Client).filter(Client.id == discovery.client_id).first()
        client_name = client.name if client else "Client"

        # Gather KPI data
        kpi_rows = []
        parameters = db.query(Parameter).filter(Parameter.process_id == process.id).all()

        for param in parameters:
            kpis = db.query(KPI).filter(KPI.parameter_id == param.id).all()
            for kpi in kpis:
                kpi_rows.append({
                    "parameter_name": param.name,
                    "kpi_name": kpi.name,
                    "formula": kpi.formula or "Measurement specification pending",
                    "unit": kpi.unit or "",
                })

        # Build Excel
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        create_instructions_sheet(wb, client_name, process.name)
        create_data_sheet(wb, kpi_rows)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "IC-Pi_Data_Template_" + client_name.replace(" ", "_") + ".xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
            headers={"Content-Disposition": "attachment; filename=" + filename},
        )

    finally:
        db.close()
